from __future__ import annotations

import argparse
import datetime as dt
import json
import math
import os
import re
from collections import Counter, defaultdict
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Iterable, Iterator


YEAR_RE = re.compile(r"(?<!\d)(20\d{2})(?!\d)")
EXCEL_EXTS = {".xls", ".xlsx", ".xlsm"}


def _safe_str(value: Any) -> str:
    try:
        return str(value)
    except Exception:
        return repr(value)


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _infer_cell_type(value: Any) -> str:
    if _is_blank(value):
        return "missing"
    if isinstance(value, bool):
        return "bool"
    # openpyxl returns datetime/date/time objects for true Excel date cells
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return "datetime"
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return "missing"
    if isinstance(value, (int, float)) and not (isinstance(value, float) and (math.isnan(value) or math.isinf(value))):
        # Excel frequently encodes integers as floats.
        if isinstance(value, float) and value.is_integer():
            return "int"
        return "float" if isinstance(value, float) else "int"
    return "str"


def _json_safe(value: Any) -> Any:
    if _is_blank(value):
        return None
    if isinstance(value, (bool, int, str)):
        return value
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return value
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        try:
            return value.isoformat()
        except Exception:
            return _safe_str(value)
    return _safe_str(value)


def _majority_type(type_counts: Counter[str]) -> str:
    counts = dict(type_counts)
    counts.pop("missing", None)
    if not counts:
        return "unknown"
    return max(counts.items(), key=lambda kv: (kv[1], kv[0]))[0]


def _extract_year_from_parts(parts: Iterable[str]) -> str | None:
    for part in parts:
        m = YEAR_RE.search(part)
        if m:
            return m.group(1)
    return None


def _station_from_path(path: Path, dataset_root: Path) -> str:
    try:
        rel = path.relative_to(dataset_root)
        # Expect: Dataset/<Station>/...
        if len(rel.parts) >= 2:
            return rel.parts[0]
        return rel.parts[0] if rel.parts else "UNKNOWN"
    except Exception:
        return "UNKNOWN"


@dataclass
class SheetProfile:
    sheet_name: str
    header_row_index_1based: int | None
    n_columns: int
    columns: list[str]
    inferred_types: dict[str, str]
    missing_rate_estimate: dict[str, float]
    sample_rows: list[dict[str, Any]]  # first N rows after header
    notes: list[str]


@dataclass
class FileProfile:
    path: str
    size_bytes: int
    modified_iso: str
    station: str
    year: str | None
    extension: str
    engine: str | None
    sheet_names: list[str]
    sheets: list[SheetProfile]
    errors: list[str]


def _score_header_row(values: list[Any]) -> float:
    non_blank = [v for v in values if not _is_blank(v)]
    if len(non_blank) < 3:
        return -1.0
    str_count = sum(1 for v in non_blank if isinstance(v, str))
    unique_strs = len({v.strip().lower() for v in non_blank if isinstance(v, str)})
    num_count = sum(1 for v in non_blank if isinstance(v, (int, float)))
    # Favor rows that look like label rows: mostly strings, many unique, few numerics.
    return (str_count / len(non_blank)) * 2.0 + (unique_strs / max(1, len(non_blank))) * 1.0 - (num_count / len(non_blank)) * 1.5


def _normalize_header(value: Any, fallback: str) -> str:
    if _is_blank(value):
        return fallback
    header = _safe_str(value).strip()
    header = re.sub(r"\s+", " ", header)
    return header if header else fallback


def _profile_sheet_from_rows(
    rows: list[list[Any]],
    sheet_name: str,
    max_sample_rows: int,
) -> SheetProfile:
    notes: list[str] = []
    if not rows:
        return SheetProfile(
            sheet_name=sheet_name,
            header_row_index_1based=None,
            n_columns=0,
            columns=[],
            inferred_types={},
            missing_rate_estimate={},
            sample_rows=[],
            notes=["Empty sheet (no readable rows)."],
        )

    # Determine the best header row among the first 20 rows.
    candidate_count = min(20, len(rows))
    best_idx = None
    best_score = -1e9
    for idx in range(candidate_count):
        score = _score_header_row(rows[idx])
        if score > best_score:
            best_score = score
            best_idx = idx

    if best_idx is None or best_score < 0:
        best_idx = 0
        notes.append("Header row heuristic weak; defaulted to first row.")

    raw_headers = rows[best_idx]
    # Clip trailing blanks
    while raw_headers and _is_blank(raw_headers[-1]):
        raw_headers = raw_headers[:-1]

    columns: list[str] = []
    seen: Counter[str] = Counter()
    for col_i, h in enumerate(raw_headers):
        base = _normalize_header(h, fallback=f"COL_{col_i+1}")
        key = base
        seen[key] += 1
        if seen[key] > 1:
            key = f"{base}__{seen[base]}"
        columns.append(key)

    if not columns:
        notes.append("No headers detected; generated placeholder columns from first data row width.")
        # Use width from first non-empty row
        width = max(len(r) for r in rows[: min(10, len(rows))])
        columns = [f"COL_{i+1}" for i in range(width)]

    data_rows = rows[best_idx + 1 : best_idx + 1 + max_sample_rows]
    # If data rows are empty, attempt to use rows after header without skipping blank rows
    if not data_rows and best_idx + 1 < len(rows):
        data_rows = rows[best_idx + 1 : min(len(rows), best_idx + 1 + max_sample_rows)]

    sample_dicts: list[dict[str, Any]] = []
    type_counts_by_col: dict[str, Counter[str]] = {c: Counter() for c in columns}
    missing_counts: Counter[str] = Counter()
    total_counts: Counter[str] = Counter()

    for r in data_rows:
        # Ensure row length
        row_values = list(r) + [None] * max(0, len(columns) - len(r))
        row_values = row_values[: len(columns)]

        row_dict: dict[str, Any] = {}
        for col, v in zip(columns, row_values):
            t = _infer_cell_type(v)
            type_counts_by_col[col][t] += 1
            total_counts[col] += 1
            if t == "missing":
                missing_counts[col] += 1
            row_dict[col] = _json_safe(v)
        sample_dicts.append(row_dict)

    inferred_types = {c: _majority_type(type_counts_by_col[c]) for c in columns}
    missing_rate = {
        c: (missing_counts[c] / total_counts[c]) if total_counts[c] else 1.0
        for c in columns
    }

    return SheetProfile(
        sheet_name=sheet_name,
        header_row_index_1based=best_idx + 1,
        n_columns=len(columns),
        columns=columns,
        inferred_types=inferred_types,
        missing_rate_estimate=missing_rate,
        sample_rows=sample_dicts,
        notes=notes,
    )


def _read_xlsx_profiles(path: Path, max_sheets: int | None, max_sample_rows: int) -> tuple[list[str], list[SheetProfile]]:
    from openpyxl import load_workbook

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        sheet_names = list(wb.sheetnames)
        target_names = sheet_names if max_sheets is None else sheet_names[:max_sheets]
        profiles: list[SheetProfile] = []
        for name in target_names:
            ws = wb[name]
            rows: list[list[Any]] = []
            # Read up to (header search rows + sample rows + a small buffer)
            limit = 20 + max_sample_rows + 5
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= limit:
                    break
                rows.append(list(row))
            profiles.append(_profile_sheet_from_rows(rows, sheet_name=name, max_sample_rows=max_sample_rows))
        return sheet_names, profiles
    finally:
        wb.close()


def _read_xls_profiles(path: Path, max_sheets: int | None, max_sample_rows: int) -> tuple[list[str], list[SheetProfile]]:
    import xlrd

    book = xlrd.open_workbook(path, on_demand=True)
    sheet_names = book.sheet_names()
    target_names = sheet_names if max_sheets is None else sheet_names[:max_sheets]
    profiles: list[SheetProfile] = []
    try:
        for name in target_names:
            sh = book.sheet_by_name(name)
            # Read up to (header search rows + sample rows + buffer)
            limit = min(sh.nrows, 20 + max_sample_rows + 5)
            rows: list[list[Any]] = []
            for r in range(limit):
                rows.append(sh.row_values(r))
            profiles.append(_profile_sheet_from_rows(rows, sheet_name=name, max_sample_rows=max_sample_rows))
    finally:
        # xlrd closes on release_resources
        book.release_resources()
    return sheet_names, profiles


def _iter_excel_files(dataset_root: Path) -> Iterator[Path]:
    for p in dataset_root.rglob("*"):
        if not p.is_file():
            continue
        if p.name.startswith("~$"):
            continue
        if p.suffix.lower() in EXCEL_EXTS:
            yield p


def _choose_sample_files(all_files: list[Path], dataset_root: Path, max_files: int) -> list[Path]:
    """
    Deterministic stratified sample: favor coverage across (station, year), then by size.
    """
    groups: dict[tuple[str, str | None], list[Path]] = defaultdict(list)
    for p in all_files:
        station = _station_from_path(p, dataset_root)
        year = _extract_year_from_parts(p.parts)
        groups[(station, year)].append(p)

    # Within each group, prefer the largest file (often the "full month"/"hourly" variant).
    for k in list(groups.keys()):
        groups[k].sort(key=lambda x: (x.stat().st_size, x.as_posix()), reverse=True)

    # First pass: one per group
    chosen: list[Path] = []
    for k in sorted(groups.keys(), key=lambda t: (t[0], t[1] or "")):
        chosen.append(groups[k][0])
        if len(chosen) >= max_files:
            return chosen

    # Second pass: fill remaining slots with next-largest files overall, avoiding duplicates.
    remaining = sorted(all_files, key=lambda x: (x.stat().st_size, x.as_posix()), reverse=True)
    chosen_set = set(chosen)
    for p in remaining:
        if p in chosen_set:
            continue
        chosen.append(p)
        chosen_set.add(p)
        if len(chosen) >= max_files:
            break
    return chosen


def _extract_pdf_text(pdf_path: Path, max_pages: int | None) -> dict[str, Any]:
    from pypdf import PdfReader

    if not pdf_path.exists():
        return {"path": str(pdf_path), "error": "PDF not found."}

    reader = PdfReader(str(pdf_path))
    pages = reader.pages
    n_pages = len(pages)
    take = n_pages if max_pages is None else min(n_pages, max_pages)
    texts: list[str] = []
    for i in range(take):
        try:
            texts.append(pages[i].extract_text() or "")
        except Exception as e:
            texts.append(f"[ERROR extracting page {i+1}: {e}]")
    return {"path": str(pdf_path), "n_pages": n_pages, "pages_extracted": take, "text": "\n\n".join(texts)}


def _schema_signature(columns: list[str]) -> str:
    # Normalize to lower, collapse spaces, remove punctuation noise
    norm = []
    for c in columns:
        v = c.strip().lower()
        v = re.sub(r"\s+", " ", v)
        v = re.sub(r"[^a-z0-9 ._%/+-]", "", v)
        norm.append(v)
    return "|".join(norm)


def _write_profile_report_md(out_path: Path, profiles: list[FileProfile]) -> None:
    # Group by schema signature per sheet name.
    schema_groups: dict[tuple[str, str], list[tuple[FileProfile, SheetProfile]]] = defaultdict(list)
    for fp in profiles:
        for sp in fp.sheets:
            sig = _schema_signature(sp.columns)
            schema_groups[(sp.sheet_name, sig)].append((fp, sp))

    lines: list[str] = []
    lines.append("# Smart Data Profiling Report (Sampled)")
    lines.append("")
    lines.append(f"- Sampled files profiled: **{len(profiles)}**")
    lines.append(f"- Distinct (sheet, schema) groups found: **{len(schema_groups)}**")
    lines.append("")

    for (sheet_name, sig), items in sorted(schema_groups.items(), key=lambda kv: (-len(kv[1]), kv[0][0])):
        fp0, sp0 = items[0]
        lines.append(f"## Sheet: {sheet_name} (n={len(items)})")
        lines.append(f"- Example file: `{fp0.path}`")
        lines.append(f"- Header row (1-based): `{sp0.header_row_index_1based}`")
        lines.append(f"- Columns ({len(sp0.columns)}):")
        for c in sp0.columns:
            t = sp0.inferred_types.get(c, "unknown")
            miss = sp0.missing_rate_estimate.get(c, 0.0)
            lines.append(f"  - `{c}`: `{t}`, missing~`{miss:.0%}`")
        if sp0.notes:
            lines.append(f"- Notes: {', '.join(sp0.notes)}")
        lines.append("")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Profile Excel dataset + extract assignment PDF text.")
    parser.add_argument("--dataset-root", type=str, default="Dataset", help="Dataset root directory (relative or absolute).")
    parser.add_argument("--pdf-path", type=str, default=r"C:\Users\ASMIT\Downloads\Machine Learning-Based Prediction of Air Pollutant Concentration.pdf")
    parser.add_argument("--out-json", type=str, default=str(Path("artifacts") / "profile_summary.json"))
    parser.add_argument("--out-md", type=str, default=str(Path("artifacts") / "profile_report.md"))
    parser.add_argument("--max-files", type=int, default=40, help="Max Excel files to fully profile (sheet+rows).")
    parser.add_argument("--max-sheets", type=int, default=3, help="Max sheets per workbook to profile (first N). Use 0 for all.")
    parser.add_argument("--sample-rows", type=int, default=50, help="Rows to sample after the header.")
    parser.add_argument("--pdf-max-pages", type=int, default=10, help="Max PDF pages to extract for quick parsing (0 for all).")
    args = parser.parse_args()

    dataset_root = Path(args.dataset_root).resolve()
    pdf_path = Path(args.pdf_path)

    all_excel = list(_iter_excel_files(dataset_root))
    all_excel_sorted = sorted(all_excel, key=lambda p: p.as_posix())
    sampled = _choose_sample_files(all_excel_sorted, dataset_root, max_files=max(1, args.max_files))

    max_sheets = None if args.max_sheets == 0 else args.max_sheets
    max_pages = None if args.pdf_max_pages == 0 else args.pdf_max_pages

    file_summaries: list[dict[str, Any]] = []
    counts_by_station: Counter[str] = Counter()
    counts_by_year: Counter[str] = Counter()
    for p in all_excel_sorted:
        station = _station_from_path(p, dataset_root)
        year = _extract_year_from_parts(p.parts) or "UNKNOWN"
        counts_by_station[station] += 1
        counts_by_year[year] += 1
        file_summaries.append(
            {
                "path": str(p),
                "size_bytes": p.stat().st_size,
                "modified_iso": dt.datetime.fromtimestamp(p.stat().st_mtime).isoformat(timespec="seconds"),
                "station": station,
                "year": year,
                "extension": p.suffix.lower(),
            }
        )

    profiles: list[FileProfile] = []
    for p in sampled:
        errors: list[str] = []
        engine: str | None = None
        sheet_names: list[str] = []
        sheets: list[SheetProfile] = []
        try:
            if p.suffix.lower() in {".xlsx", ".xlsm"}:
                engine = "openpyxl"
                sheet_names, sheets = _read_xlsx_profiles(p, max_sheets=max_sheets, max_sample_rows=args.sample_rows)
            elif p.suffix.lower() == ".xls":
                engine = "xlrd"
                sheet_names, sheets = _read_xls_profiles(p, max_sheets=max_sheets, max_sample_rows=args.sample_rows)
            else:
                errors.append(f"Unsupported extension: {p.suffix}")
        except Exception as e:
            errors.append(f"{type(e).__name__}: {_safe_str(e)}")

        station = _station_from_path(p, dataset_root)
        year = _extract_year_from_parts(p.parts)
        profiles.append(
            FileProfile(
                path=str(p),
                size_bytes=p.stat().st_size,
                modified_iso=dt.datetime.fromtimestamp(p.stat().st_mtime).isoformat(timespec="seconds"),
                station=station,
                year=year,
                extension=p.suffix.lower(),
                engine=engine,
                sheet_names=sheet_names,
                sheets=sheets,
                errors=errors,
            )
        )

    pdf_extract = _extract_pdf_text(pdf_path, max_pages=max_pages)

    out_json = Path(args.out_json)
    out_json.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "generated_at_iso": dt.datetime.now().isoformat(timespec="seconds"),
        "dataset_root": str(dataset_root),
        "excel_files_total": len(all_excel_sorted),
        "excel_files_profiled": len(sampled),
        "profile_settings": {
            "max_files": args.max_files,
            "max_sheets": args.max_sheets,
            "sample_rows": args.sample_rows,
        },
        "counts_by_station": dict(counts_by_station),
        "counts_by_year": dict(counts_by_year),
        "all_excel_files": file_summaries,
        "sampled_profiles": [asdict(p) for p in profiles],
        "assignment_pdf": pdf_extract,
    }
    out_json.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")

    _write_profile_report_md(Path(args.out_md), profiles=profiles)

    # Print a tiny summary for CLI use.
    print(f"Excel files total: {len(all_excel_sorted)}")
    print(f"Excel files profiled (sample): {len(sampled)}")
    print(f"Report: {Path(args.out_md).resolve()}")
    print(f"JSON: {out_json.resolve()}")
    if isinstance(pdf_extract, dict) and pdf_extract.get("error"):
        print(f"PDF: error: {pdf_extract.get('error')}")
    else:
        print(f"PDF pages extracted: {pdf_extract.get('pages_extracted')} / {pdf_extract.get('n_pages')}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
